# auto_input_simpus_final_SDN104214_fix_regions_kel.py
# -----------------------------------------------------------
# Gaya lama (openpyxl + row[index]) + tanggal mm/dd/YYYY
# + mapping Provinsi & Kab/Kota dari List_Kecamatan.xlsx
# + isi Kelurahan dari kolom Excel siswa (tanpa perlu mapping baru)
# -----------------------------------------------------------

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import datetime
import time, re, sys, unicodedata
import pandas as pd

# === KONFIGURASI ===
URL_LOGIN = "https://dinkesds-simpus.deliserdangkab.go.id/"
URL_PASIEN = "https://dinkesds-simpus.deliserdangkab.go.id/pasien"
EXCEL_PATH_SISWA = "SDN 104214 1-6 2025.xlsx"
EXCEL_PATH_KEC   = "List_Kecamatan.xlsx"

# === INPUT USER ===
NOMOR_SISWA_AWAL = int(input("Masukkan nomor urut siswa awal (1-based): "))
NOMOR_SISWA_AKHIR = int(input("Masukkan nomor urut siswa akhir: "))
if NOMOR_SISWA_AKHIR < NOMOR_SISWA_AWAL:
    raise ValueError("❌ Nomor siswa akhir harus >= awal.")

# ===== UTIL =====
def _norm_text(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = s.replace("KECAMATAN", " ").replace("KEC.", " ").replace("KEC ", " ")
    s = s.replace("KABUPATEN", " ").replace("KAB.", " ").replace("KAB ", " ")
    s = s.replace("KOTA ADMINISTRASI", "KOTA")
    s = s.replace("-", " ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _compact(s: str) -> str:
    return _norm_text(s).replace(" ", "")

def fmt_mmddyyyy(val):
    """yyyy-mm-dd / variasi lain / datetime -> mm/dd/YYYY; gagal -> ''."""
    if val is None: return ""
    try:
        if isinstance(val, datetime):
            return val.strftime("%m/%d/%Y")
        s = str(val).strip()
        for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%m/%d/%Y","%Y/%m/%d","%d.%m.%Y"):
            try:
                d = datetime.strptime(s, fmt)
                return d.strftime("%m/%d/%Y")
            except Exception:
                continue
        d = pd.to_datetime(s, dayfirst=True, errors="raise")
        return d.strftime("%m/%d/%Y")
    except Exception:
        return ""

def clean_kecamatan_input(kec_raw: str) -> str:
    k = str(kec_raw or "").strip()
    if not k: return ""
    low = k.lower()
    if low in ["stm hilir","s.t.m hilir","st m hilir","stmhilir"]:
        return "SINEMBAH TANJUNG MUDA HILIR"
    if low in ["stm hulu","s.t.m hulu","st m hulu","stmhulu"]:
        return "SINEMBAH TANJUNG MUDA HULU"
    up = _norm_text(k)
    if "SIBIRU BIRU" in up or "SI BIRU" in up:
        up = "BIRU BIRU"
    return up

def clean_kelurahan_input(vill_raw: str) -> str:
    """Bersihkan nilai kelurahan: buang 'Desa/Kel.', 'Kel.', 'Desa', dsb."""
    s = str(vill_raw or "").strip()
    if not s: return ""
    up = _norm_text(s)
    # hapus awalan umum
    up = up.replace("DESA KEL", " ").replace("DESA/KEL", " ")
    up = up.replace("DESA", " ").replace("KELURAHAN", " ").replace("KEL", " ").replace("KEL.", " ")
    up = re.sub(r"\s+", " ", up).strip()
    return up

# ===== BACA REFERENSI List_Kecamatan.xlsx =====
def _find_col(df, keywords):
    for c in df.columns:
        n = _norm_text(c)
        if any(kw in n for kw in keywords):
            return c
    return None

try:
    df_ref = pd.read_excel(EXCEL_PATH_KEC, sheet_name=0, dtype=str)
except Exception as e:
    print(f"❌ Gagal baca {EXCEL_PATH_KEC}: {e}")
    sys.exit(1)

col_kec = _find_col(df_ref, ["KEC"]) or _find_col(df_ref, ["KECAMATAN"])
col_kab = _find_col(df_ref, ["KAB", "KOTA"])
col_prov= _find_col(df_ref, ["PROV"])
if not all([col_kec, col_kab, col_prov]):
    print("❌ List_Kecamatan.xlsx harus punya kolom berisi 'Kecamatan', 'Kabupaten/Kota', dan 'Provinsi'.")
    sys.exit(1)

kec_to_region = {}
for _, r in df_ref.iterrows():
    kec_val  = (r.get(col_kec)  or "").strip()
    kab_val  = (r.get(col_kab)  or "").strip()
    prov_val = (r.get(col_prov) or "").strip()
    if not kec_val:
        continue
    base = _norm_text(kec_val)
    for k in {base, _compact(kec_val), base.replace(" ", "")}:
        kec_to_region[k] = (kab_val, prov_val)

def lookup_region(kec_raw: str):
    if not kec_raw:
        return ("", "")
    normalized = clean_kecamatan_input(kec_raw)
    keys = [normalized, normalized.replace(" ", ""), _compact(kec_raw)]
    for k in keys:
        if k in kec_to_region:
            return kec_to_region[k]
    for k in kec_to_region:
        if k in normalized or normalized in k:
            return kec_to_region[k]
    return ("", "")

# ====== DETEKSI HEADER FILE SISWA + INDEX KOLUM ======
wb = load_workbook(EXCEL_PATH_SISWA)
sheet = wb.active

def _norm_cell(v): return re.sub(r"\s+"," ",str(v or "").strip()).lower()

header_row, header_vals = None, None
for r in range(1, 16):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column+1)]
    n = [_norm_cell(v) for v in vals]
    if any("nama" in x for x in n) and (any(x=="jk" for x in n) or any("l p" in x for x in n) or any("l/p" in x for x in n)) and any("tanggal lahir" in x for x in n):
        header_row, header_vals = r, vals
        break
if header_row is None:
    header_row, header_vals = 1, [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column+1)]

print(f"ℹ️ Header terdeteksi di baris Excel: {header_row}")

name_to_idx = {_norm_cell(h): i for i, h in enumerate(header_vals)}

def find_col(*aliases):
    for a in aliases:
        a = _norm_cell(a)
        if a in name_to_idx: return name_to_idx[a]
    for a in aliases:
        a = _norm_cell(a)
        for k,i in name_to_idx.items():
            if a in k: return i
    return None

IDX_NAMA      = find_col("nama","nama siswa","nama peserta","nama lengkap")
IDX_JK        = find_col("jk","l p","l/p","jenis kelamin","kelamin")
IDX_TMP_LAHIR = find_col("tempat lahir","tmpt lahir","kota lahir")
IDX_TGL_LAHIR = find_col("tanggal lahir","tgl lahir","lahir")
IDX_NIK       = find_col("nik","no ktp","nomor ktp","no. ktp","no nik","no. nik")
IDX_AGAMA     = find_col("agama")
IDX_ALAMAT    = find_col("alamat","alamat domisili","alamat rumah")
IDX_KEC       = find_col("kecamatan","kec","kecamatan domisili")
# >>> Tambah deteksi kolom kelurahan <<<
IDX_KEL       = find_col("kelurahan","desa","desa/kel","desa/kelurahan","desa/kel.")

print("ℹ️ Index kolom:", {
    "nama":IDX_NAMA,"jk":IDX_JK,"tempat_lahir":IDX_TMP_LAHIR,"tanggal_lahir":IDX_TGL_LAHIR,
    "nik":IDX_NIK,"agama":IDX_AGAMA,"alamat":IDX_ALAMAT,"kecamatan":IDX_KEC,"kelurahan":IDX_KEL
})

DATA_FIRST_ROW = header_row + 1

# ====== BROWSER ======
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
driver.get(URL_LOGIN)
print("✅ Silakan login manual (captcha) lalu klik LOGIN")
input("🔐 Tekan ENTER setelah berhasil login...")

driver.get(URL_PASIEN)
time.sleep(2)

# ====== LOOP ======
JUMLAH_DATA = NOMOR_SISWA_AKHIR - NOMOR_SISWA_AWAL + 1

for i in range(JUMLAH_DATA):
    excel_row = DATA_FIRST_ROW + (NOMOR_SISWA_AWAL - 1) + i
    row = sheet[excel_row]

    def get(idx): return None if idx is None else row[idx].value

    nama         = get(IDX_NAMA)
    if not nama:
        print(f"⚠️ Baris {excel_row}: nama kosong. Skip."); 
        continue
    jk           = str(get(IDX_JK) or "").strip()
    tempat_lahir = get(IDX_TMP_LAHIR)
    tgl_raw      = get(IDX_TGL_LAHIR)
    nik          = str(get(IDX_NIK) or "")
    agama        = str(get(IDX_AGAMA) or "").strip()
    alamat       = get(IDX_ALAMAT)
    kec_raw      = get(IDX_KEC)
    kel_raw      = get(IDX_KEL)

    tanggal_lahir = fmt_mmddyyyy(tgl_raw)
    if not tanggal_lahir:
        print(f"⚠️ Gagal format tanggal: {nama} -> {tgl_raw}")

    kabupaten_disp, provinsi_disp = lookup_region(kec_raw)
    kelurahan_clean = clean_kelurahan_input(kel_raw)

    print(f"\n🟢 Siapkan form untuk: {nama} | Kec='{kec_raw}' → Kab='{kabupaten_disp}' | Prov='{provinsi_disp}' | Kel='{kelurahan_clean}'")
    input("➡️ Setelah klik 'TAMBAH DATA' & dropdown wilayah muncul, tekan ENTER untuk isi otomatis...")

    try:
        # === Identitas dasar ===
        driver.find_element(By.XPATH, '//input[@placeholder="Nama"]').clear()
        driver.find_element(By.XPATH, '//input[@placeholder="Nama"]').send_keys(str(nama))

        Select(driver.find_element(By.XPATH, '//select[./option[contains(text(), "Jenis Kelamin")]]')) \
            .select_by_visible_text("Perempuan" if jk.upper() == "P" else "Laki-Laki")

        driver.find_element(By.XPATH, '//input[@placeholder="Tempat Lahir"]').clear()
        driver.find_element(By.XPATH, '//input[@placeholder="Tempat Lahir"]').send_keys(str(tempat_lahir or ""))

        el = driver.find_element(By.XPATH, '//input[@placeholder="Tanggal Lahir"]')
        el.clear(); el.send_keys(tanggal_lahir)

        driver.find_element(By.XPATH, '//input[@placeholder="NIK"]').clear()
        driver.find_element(By.XPATH, '//input[@placeholder="NIK"]').send_keys(nik)

        if agama:
            Select(driver.find_element(By.XPATH, '//select[./option[contains(text(), "Agama")]]')) \
                .select_by_visible_text(agama.upper())

        driver.find_element(By.XPATH, '//input[@placeholder="Alamat"]').clear()
        driver.find_element(By.XPATH, '//input[@placeholder="Alamat"]').send_keys(str(alamat or ""))

        # === Provinsi ===
        if provinsi_disp:
            prov_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//label[contains(text(), "Provinsi")]/following-sibling::div//input'))
            )
            prov_input.click(); time.sleep(0.3); prov_input.clear()
            prov_input.send_keys(provinsi_disp); time.sleep(1.0)
            prov_input.send_keys(Keys.ARROW_DOWN); prov_input.send_keys(Keys.ENTER)

        # === Kabupaten/Kota ===
        if kabupaten_disp:
            kab_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//label[contains(text(), "Kabupaten")]/following-sibling::div//input'))
            )
            kab_input.click(); time.sleep(0.3); kab_input.clear()
            kab_to_type = kabupaten_disp.strip()
            if not kab_to_type.upper().startswith(("KABUPATEN ", "KOTA ")):
                kab_to_type = "KABUPATEN " + kab_to_type
            kab_input.send_keys(kab_to_type); time.sleep(1.0)
            kab_input.send_keys(Keys.ARROW_DOWN); kab_input.send_keys(Keys.ENTER)

        # === Kecamatan ===
        if kec_raw:
            kec_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//label[contains(text(), "Kecamatan")]/following-sibling::div//input'))
            )
            kec_input.click(); time.sleep(0.3); kec_input.clear()
            kec_try = clean_kecamatan_input(kec_raw)
            kec_input.send_keys(kec_try); time.sleep(0.8)
            kec_input.send_keys(Keys.ARROW_DOWN); kec_input.send_keys(Keys.ENTER)

        # === Kelurahan ===
        if kelurahan_clean:
            # tunggu data kelurahan ter-load setelah pilih kecamatan
            time.sleep(1.0)
            kel_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//label[contains(text(), "Kelurahan")]/following-sibling::div//input'))
            )
            kel_input.click(); time.sleep(0.3); kel_input.clear()
            # coba ketik versi bersih
            kel_input.send_keys(kelurahan_clean); time.sleep(0.9)
            kel_input.send_keys(Keys.ARROW_DOWN); kel_input.send_keys(Keys.ENTER)
        else:
            print(f"⚠️ Kelurahan kosong/tdk dikenali untuk {nama}. Lewati pengisian kelurahan.")

        print("✅ Data berhasil diisi. Silakan klik tombol TAMBAH.")
    except Exception as e:
        print(f"❌ Gagal input untuk {nama} → {e}")
        continue

print("\n🍀 Selesai.")
