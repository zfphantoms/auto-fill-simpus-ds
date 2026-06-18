from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
import sys
import os
import tkinter as tk
from tkinter import filedialog

# Explicit imports to help PyInstaller bundle dynamic dependencies
import selenium.webdriver.chrome.webdriver
import selenium.webdriver.chrome.service
import selenium.webdriver.chrome.options

# Membuka browser Chrome (via built-in Selenium Manager)
options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)  # Agar browser tidak langsung tertutup

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 20)

# URLs
login_url = 'https://sipp.bpjs-kesehatan.go.id/sipp/#/access/signin'
dashboard_url_fragment = '#/app/dashboardadmin'
pencarian_url = 'https://sipp.bpjs-kesehatan.go.id/sipp/#/app/pencarian'

# Input/Output files
output_xlsx = 'hasil_peserta_sipp.xlsx'

# === PILIH FILE EXCEL (via File Explorer) ===
_root = tk.Tk()
_root.withdraw()
_root.attributes('-topmost', True)

print("\n📂 Silakan pilih file Excel di jendela yang muncul...")
input_xlsx = filedialog.askopenfilename(
    title="Pilih file Excel data SIPP",
    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    initialdir="."
)
_root.destroy()

if not input_xlsx:
    print("❌ Tidak ada file yang dipilih. Keluar.")
    sys.exit(1)

print(f"✅ File yang dipilih: {input_xlsx}\n")

# Konversi format .xls ke .xlsx jika diperlukan
if input_xlsx.lower().endswith('.xls'):
    print("ℹ️ Format .xls terdeteksi. Mengonversi ke format .xlsx...")
    try:
        import xlrd
    except ImportError:
        print("❌ Library 'xlrd' belum terpasang. Library ini dibutuhkan untuk membaca file .xls.")
        print("ℹ️ Menginstal 'xlrd' secara otomatis...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd"])
            import xlrd
            print("✅ 'xlrd' berhasil dipasang.")
        except Exception as inst_err:
            print(f"❌ Gagal memasang 'xlrd' secara otomatis: {inst_err}")
            print("Silakan jalankan perintah berikut secara manual terlebih dahulu di terminal:")
            print("   pip install xlrd")
            sys.exit(1)
    
    try:
        TEMP_XLSX_PATH = os.path.splitext(input_xlsx)[0] + "_temp_converted.xlsx"
        print(f"ℹ️ Mengonversi {input_xlsx} -> {TEMP_XLSX_PATH} ...")
        xls_data = pd.read_excel(input_xlsx, sheet_name=None, header=None)
        with pd.ExcelWriter(TEMP_XLSX_PATH, engine='openpyxl') as writer:
            for sheet_name, df_sheet in xls_data.items():
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        input_xlsx = TEMP_XLSX_PATH
        print("✅ Konversi ke .xlsx selesai.")
    except Exception as conv_err:
        print(f"❌ Gagal mengonversi file .xls ke .xlsx: {conv_err}")
        sys.exit(1)

# === INPUT USER ===
try:
    NOMOR_AWAL = int(input("Masukkan angka baris data awal yang mau di-get (contoh 1): "))
    NOMOR_AKHIR = int(input("Masukkan angka baris data akhir yang mau di-get (contoh 1000): "))
    if NOMOR_AKHIR < NOMOR_AWAL:
        raise ValueError("❌ Nomor akhir harus >= awal.")
except ValueError as ve:
    print(f"Input tidak valid: {ve}")
    sys.exit(1)

# Variasi nama kolom untuk deteksi dinamis (diambil dari auto_input_simpus_fix)
NIK_ALIASES = ["NIK", "NO KTP", "NOMOR KTP", "NO. KTP", "NO NIK", "NO. NIK", "NIK/KITAS/KITAP", "NOMOR INDUK KEPENDUDUKAN"]
NAMA_ALIASES = ["NAMA ANGGOTA KELUARGA", "NAMA", "NAMA SISWA", "NAMA PESERTA", "NAMA LENGKAP"]
TGL_LAHIR_ALIASES = ["TANGGAL LAHIR", "TGL LAHIR", "LAHIR"]

# Kolom tambahan dari SIPP (ditambah di akhir)
SIPP_COLUMNS = [
    'SIPP_Nomor Kartu',
    'SIPP_Nama',
    'SIPP_Status Kepesertaan',
    'SIPP_Hak Kelas Rawat',
    'SIPP_Segmen Peserta',
    'SIPP_FKTP Terdaftar',
    'SIPP_No. VA Bank Mandiri',
    'SIPP_No. VA Non Bank Mandiri',
]


def normalize_col(name: object) -> str:
    """Samakan nama kolom: UPPERCASE, hilangkan newline dan spasi berlebih."""
    if name is None: return ""
    s = str(name).upper()
    s = s.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
    s = ' '.join(s.split())  # collapse whitespace
    return s.strip()

def find_column_key(columns, aliases):
    """Cari nama kolom berdasarkan alias (exact match dulu, lalu partial match)."""
    # Exact match
    for a in aliases:
        a_norm = normalize_col(a)
        if a_norm in columns:
            return a_norm
    # Partial match
    for a in aliases:
        a_norm = normalize_col(a)
        for c in columns:
            if a_norm in c:
                return c
    return None


def create_output_workbook(path: str, output_columns: list[str]) -> None:
    """Buat file Excel hasil dari awal (overwrite) + header sesuai kolom output."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'hasil'
    ws.append(output_columns)
    wb.save(path)


def append_row_to_output(path: str, output_columns: list[str], row: dict) -> None:
    """Append 1 baris hasil ke Excel tanpa menulis ulang seluruh file."""
    wb = load_workbook(path)
    ws = wb.active
    ws.append([row.get(c) for c in output_columns])
    wb.save(path)


def wait_until_logged_in(timeout_seconds: int = 300) -> None:
    """Tunggu sampai user selesai login (terdeteksi dari URL dashboard)."""
    start = time.time()
    while True:
        current_url = driver.current_url or ''
        if dashboard_url_fragment in current_url:
            return
        if time.time() - start > timeout_seconds:
            raise TimeoutError('Timeout: belum terdeteksi login (belum masuk dashboard).')
        time.sleep(1)


def open_pencarian() -> None:
    driver.get(pencarian_url)
    # Tunggu container halaman pencarian muncul
    wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(.,'Pencarian Peserta')]")))


def select_mode_nik() -> None:
    """Klik tab/opsi 'NIK' pada Pencarian Detail Peserta."""
    nik_label = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                "//label[(normalize-space()='NIK' or contains(normalize-space(.),'NIK')) and (@ng-click='changeJenis(1)' or contains(@ng-click,'changeJenis'))]",
            )
        )
    )
    nik_label.click()


def find_search_input() -> 'webdriver.remote.webelement.WebElement':
    """Cari textbox input pencarian pada panel 'Pencarian Detail Peserta'."""
    candidates = driver.find_elements(By.XPATH, "//input[(@type='text' or not(@type)) and not(@disabled)]")
    visibles = [c for c in candidates if c.is_displayed()]
    if not visibles:
        raise RuntimeError('Tidak menemukan textbox input pencarian.')
    for c in visibles:
        ph = (c.get_attribute('placeholder') or '').lower()
        if 'cari' in ph or 'nik' in ph or 'nokap' in ph:
            return c
    return visibles[0]


def scrape_detail_peserta() -> dict:
    """Ambil data dari tabel Detail Peserta (kiri) setelah hasil tampil."""
    wait.until(
        EC.presence_of_element_located(
            (By.XPATH, "//*[self::td or self::th][contains(normalize-space(.),'Nomor Kartu')]")
        )
    )

    def get_value(label: str):
        xpath = (
            "(//tr[td[1][contains(normalize-space(.),'{label}')]]/td[2] | "
            "//tr[th[1][contains(normalize-space(.),'{label}')]]/td[1])[1]"
        ).format(label=label)
        els = driver.find_elements(By.XPATH, xpath)
        if not els:
            els = driver.find_elements(
                By.XPATH,
                f"(//*[self::td or self::th][contains(normalize-space(.),'{label}')]/following::td[1])[1]",
            )
        if not els:
            return None
        txt = (els[0].text or '').strip()
        return txt if txt != '' else None

    return {
        'SIPP_Nomor Kartu': get_value('Nomor Kartu'),
        'SIPP_Nama': get_value('Nama'),
        'SIPP_Status Kepesertaan': get_value('Status') or get_value('Status Kepesertaan'),
        'SIPP_Hak Kelas Rawat': get_value('Hak Kelas'),
        'SIPP_Segmen Peserta': get_value('Segmen'),
        'SIPP_FKTP Terdaftar': get_value('FKTP'),
        'SIPP_No. VA Bank Mandiri': get_value('No. VA Bank'),
        'SIPP_No. VA Non Bank Mandiri': get_value('No. VA Non'),
    }


def read_input_rows(path: str) -> tuple[pd.DataFrame, str, str, str]:
    df = pd.read_excel(path, dtype=str)
    df.columns = [normalize_col(c) for c in df.columns]

    nik_col = find_column_key(df.columns, NIK_ALIASES)
    if not nik_col:
        raise ValueError(
            'Kolom NIK tidak ditemukan. Pastikan ada kolom dengan kata kunci: '
            + ', '.join(NIK_ALIASES)
        )

    nama_col = find_column_key(df.columns, NAMA_ALIASES)
    tgl_lahir_col = find_column_key(df.columns, TGL_LAHIR_ALIASES)

    return df, nik_col, nama_col, tgl_lahir_col


def normalize_date_only(value: object) -> str | None:
    """Ubah nilai tanggal menjadi string YYYY-MM-DD (tanpa jam)."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() == 'nan':
        return None
    # pandas sering baca tanggal jadi 'YYYY-MM-DD HH:MM:SS'
    if len(s) >= 10:
        return s[:10]
    return s


print('Membuka halaman login...')
driver.get(login_url)
print('Silakan login manual di browser. Script akan lanjut otomatis saat sudah masuk dashboard.')

try:
    wait_until_logged_in(timeout_seconds=300)
    open_pencarian()
    select_mode_nik()

    input_df, nik_col, nama_col, tgl_lahir_col = read_input_rows(input_xlsx)
    input_columns = list(input_df.columns)
    output_columns = input_columns + [c for c in SIPP_COLUMNS if c not in input_columns]

    # Potong dataframe berdasarkan input user
    # NOMOR_AWAL adalah 1-based index (data ke-1 berarti index ke-0)
    start_idx = max(0, NOMOR_AWAL - 1)
    end_idx = min(len(input_df), NOMOR_AKHIR)
    sliced_df = input_df.iloc[start_idx:end_idx]

    # Buat file hasil dari awal, kolom sama dengan input + kolom SIPP di akhir
    create_output_workbook(output_xlsx, output_columns)

    search_input = find_search_input()

    total_diproses = len(sliced_df)
    for i, (original_idx, row) in enumerate(sliced_df.iterrows()):
        raw_nik = str(row.get(nik_col) or '').strip()
        if raw_nik.lower() == 'nan':
            raw_nik = ''
            
        # hilangkan spasi dan .0 (kasus pandas read excel number)
        nik = raw_nik.replace(' ', '')
        if nik.endswith('.0'):
            nik = nik[:-2]
            
        # Ekstrak hanya digit (angka) untuk menangani variasi teks seperti 'tidak ada', '-', dll
        nik = ''.join(ch for ch in nik if ch.isdigit())

        # Siapkan row output dari row input
        out = {col: (row.get(col) if col in row else None) for col in input_columns}

        # Normalisasi kolom tanggal agar hanya tanggal saja
        if tgl_lahir_col and tgl_lahir_col in out:
            out[tgl_lahir_col] = normalize_date_only(out.get(tgl_lahir_col))

        if not nik or nik.lower() == 'nan':
            # tetap tulis baris walau NIK kosong
            append_row_to_output(output_xlsx, output_columns, out)
            print(f'[{i+1}/{total_diproses}] (Data ke-{original_idx+1}) SKIP (NIK kosong)')
            with open("log-failed-sipp.txt", "a") as f:
                f.write(f"Data ke-{original_idx+1} | NIK kosong\n")
            continue

        # Hapus notifikasi (toaster) yang mungkin menghalangi klik
        driver.execute_script("""
            var toasts = document.querySelectorAll('.toast-error, .toast');
            for(var i=0; i<toasts.length; i++){
                toasts[i].remove();
            }
        """)

        # Input NIK dan ENTER
        try:
            search_input.click()
        except Exception:
            driver.execute_script("arguments[0].click();", search_input)
            
        search_input.send_keys(Keys.CONTROL, 'a')
        search_input.send_keys(Keys.BACKSPACE)
        search_input.send_keys(nik)
        search_input.send_keys(Keys.ENTER)

        time.sleep(1)

        try:
            sipp_data = scrape_detail_peserta()
            out.update(sipp_data)

            # Jika nama dari SIPP '-' atau kosong, pakai nama dari tabel awal
            sipp_name = (out.get('SIPP_Nama') or '').strip()
            input_name = (out.get(nama_col) or '').strip() if nama_col else ''
            if (not sipp_name) or sipp_name == '-':
                if input_name:
                    out['SIPP_Nama'] = input_name
                    if nama_col:
                        out[nama_col] = input_name

            # Update kolom nama dari SIPP (jika kolomnya ada di input) - hanya jika SIPP memberi nama.
            elif nama_col and nama_col in out:
                out[nama_col] = out.get('SIPP_Nama')

            append_row_to_output(output_xlsx, output_columns, out)
            print(f'[{i+1}/{total_diproses}] (Data ke-{original_idx+1}) OK: {nik} -> append ke {output_xlsx}')
            
            # Tulis NIK terakhir yang berhasil
            with open("last-input-sipp.txt", "w") as f:
                f.write(f"Data ke-{original_idx+1} | NIK: {nik}")

        except Exception as e:
            print(f'[{i+1}/{total_diproses}] (Data ke-{original_idx+1}) GAGAL: {nik} -> SIPP Timeout / Tidak Ditemukan')
            append_row_to_output(output_xlsx, output_columns, out) # Tetap tulis data inputnya tanpa hasil SIPP
            with open("log-failed-sipp.txt", "a") as f:
                f.write(f"Data ke-{original_idx+1} | NIK: {nik} | Error: SIPP Timeout / Tidak Ditemukan\n")

    print(f'Selesai. Memproses {total_diproses} data. File hasil disimpan ke: {output_xlsx}')

    print('Tekan Ctrl+C untuk menutup browser...')
    while True:
        time.sleep(1)
except KeyboardInterrupt:
    driver.quit()
except Exception as e:
    print(f'Error: {e}')
    driver.quit()
