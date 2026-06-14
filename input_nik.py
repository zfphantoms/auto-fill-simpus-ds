from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import pandas as pd

# Konfigurasi opsi Chrome
options = Options()
options.add_experimental_option("detach", True)  # Agar browser tidak langsung tertutup

# Inisialisasi driver Chrome
# Pastikan chromedriver.exe ada di folder yang sama dengan script ini

driver = webdriver.Chrome(options=options)
driver.maximize_window()

# Buka website BPJS SIPP
driver.get("https://sipp.bpjs-kesehatan.go.id/sipp/#/access/signin")

print("Website BPJS SIPP sudah dibuka.")
print("Silakan login secara manual.")
print("Username untuk login: 02042701L01")
print("Password untuk login: Malam*123")
print("Copy-paste username dan password di halaman login.")
time.sleep(10)

# Tunggu sampai user login dan masuk ke dashboard
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

try:
    WebDriverWait(driver, 600).until(
        lambda d: d.current_url.startswith("https://sipp.bpjs-kesehatan.go.id/sipp/#/app/dashboardadmin")
    )
    print("Berhasil login. Silakan lanjut ke halaman pencarian secara manual.")
    # Tunggu sampai user masuk ke halaman pencarian
    WebDriverWait(driver, 600).until(
        lambda d: d.current_url.startswith("https://sipp.bpjs-kesehatan.go.id/sipp/#/app/pencarian")
    )
    print("Berhasil masuk ke halaman pencarian. Siap lanjut proses berikutnya.")
    print("Mohon pastikan Anda sudah mengklik tombol 'NIK' (bukan 'NOKAPST') di form 'Pencarian Detail Peserta' sebelum melanjutkan.")
    input("Tekan ENTER jika sudah memilih 'NIK'...")
except Exception as e:
    print(f"Gagal mendeteksi halaman dashboard atau pencarian: {e}")

# Setelah user konfirmasi sudah memilih 'NIK', lanjut input NIK dari hasil_nik.xlsx
try:
    df_nik = pd.read_excel('hasil_nik.xlsx')
    nik_list = df_nik['NIK'].dropna().astype(str).tolist()
    print(f"Total NIK yang akan diinput: {len(nik_list)}")
    for idx, nik in enumerate(nik_list, 1):
        try:
            nik_input = driver.find_element(By.XPATH, '//input[@ng-model="cari" and @type="text"]')
            nik_input.clear()
            nik_input.send_keys(nik)
            print(f"NIK ke-{idx}: {nik} sudah diinput. Silakan klik 'Cari' dan cek hasil.")
            input("Tekan ENTER untuk lanjut ke NIK berikutnya...")
            # Setelah user klik 'Cari', ambil data dari tabel Detail Peserta
            try:
                detail_rows = driver.find_elements(By.XPATH, '//div[contains(@class,"panel-body")]//table//tr')
                data = {}
                for row in detail_rows:
                    ths = row.find_elements(By.TAG_NAME, 'th')
                    tds = row.find_elements(By.TAG_NAME, 'td')
                    if len(ths) == 1 and len(tds) == 1:
                        key = ths[0].text.strip()
                        value = tds[0].text.strip()
                        data[key] = value
                    elif len(tds) == 2:
                        key = tds[0].text.strip()
                        value = tds[1].text.strip()
                        data[key] = value
                data['NIK'] = nik
                # Urutan kolom tetap
                columns = [
                    'NIK', 'Nomor Kartu', 'Nama', 'Status Kepesertaan', 'Hak Kelas Rawat',
                    'Segmen Peserta', 'FKTP Terdaftar', 'No. VA Bank Mandiri', 'No. VA Non Bank Mandiri'
                ]
                # Susun data sesuai urutan kolom
                row_data = [data.get(col, '') for col in columns]
                import os
                from openpyxl import Workbook, load_workbook
                file_excel = 'hasil_nik_sipp.xlsx'
                if not os.path.exists(file_excel):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(columns)
                    wb.save(file_excel)
                wb = load_workbook(file_excel)
                ws = wb.active
                ws.append(row_data)
                wb.save(file_excel)
                print(f"Data peserta NIK {nik} berhasil ditulis ke hasil_nik_sipp.xlsx.")
            except Exception as e:
                print(f"Gagal mengambil/menulis data peserta NIK {nik}: {e}")
        except Exception as e:
            print(f"Gagal input NIK ke-{idx}: {nik} → {e}")
except Exception as e:
    print(f"Gagal membaca file hasil_nik.xlsx atau input NIK: {e}")

# (Browser tetap terbuka karena detach=True)
