# simpus.py
# Mode MANUAL:
# - Kamu klik "TAMBAH DATA" sendiri (script menunggu ENTER lalu isi otomatis)
# - Kamu juga klik "TAMBAH" sendiri
# Fitur:
# - Tanggal lahir -> PILIH: mm/dd/YYYY atau dd/mm/YYYY saat runtime
# - Prov/Kab/Kec dari List_Kecamatan.xlsx (robust)
# - Kelurahan dari Excel
# - Log gagal ke failed-log.txt
# - Khusus SIBIRU-BIRU: yang DIKETIK ke UI = "BIRU" (mapping tetap "BIRU-BIRU")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook
from datetime import datetime
import time, re, sys, unicodedata
import pandas as pd

# === KONFIGURASI ===
URL_LOGIN = "https://dinkesds-simpus.deliserdangkab.go.id/"
URL_PASIEN = "https://dinkesds-simpus.deliserdangkab.go.id/pasien"
EXCEL_PATH_SISWA = "DATA SDN 104213 2025.xlsx"
EXCEL_PATH_KEC   = "List_Kecamatan.xlsx"
FAILED_LOG_PATH  = "failed-log.txt"

# === INPUT USER ===
NOMOR_SISWA_AWAL = int(input("Masukkan nomor urut siswa awal (1-based): "))
NOMOR_SISWA_AKHIR = int(input("Masukkan nomor urut siswa akhir: "))
if NOMOR_SISWA_AKHIR < NOMOR_SISWA_AWAL:
    raise ValueError("❌ Nomor siswa akhir harus >= awal.")

# Pilih format tanggal output
fmt_choice = input("Pilih format tanggal lahir (1=mm/dd/yyyy, 2=dd/mm/yyyy) [1]: ").strip()
OUT_DATE_FMT = "%d/%m/%Y" if fmt_choice == "2" else "%m/%d/%Y"
print(f"ℹ️ Output tanggal akan menggunakan format: {OUT_DATE_FMT.replace('%','').lower()}")

# ===== UTIL =====
def _norm_text(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = s.replace("KECAMATAN", " ").replace("KEC.", " ").replace("KEC ", " ")
    s = s.replace("KABUPATEN", " ").replace("KAB.", " ").replace("KAB ", " ")
    s = s.replace("KOTA ADMINISTRASI", "KOTA")
    s = s.replace("-", " ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _compact(s: str) -> str:
    return _norm_text(s).replace(" ", "")

def fmt_date(val, out_fmt="%m/%d/%Y"):
    """Parse berbagai input (datetime/string) lalu format ke out_fmt."""
    if val is None:
        return ""
    try:
        if isinstance(val, datetime):
            return val.strftime(out_fmt)
        s = str(val).strip()
        if not s:
            return ""
        # Coba beberapa format umum dulu
        in_fmts = (
            "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
            "%Y/%m/%d", "%d.%m.%Y", "%m-%d-%Y", "%m.%d.%Y"
        )
        for f in in_fmts:
            try:
                d = datetime.strptime(s, f)
                return d.strftime(out_fmt)
            except Exception:
                pass
        # Fallback: pandas to_datetime, coba dayfirst True lalu False
        for dayfirst in (True, False):
            d = pd.to_datetime(s, dayfirst=dayfirst, errors="coerce")
            if not pd.isna(d):
                return pd.Timestamp(d).strftime(out_fmt)
    except Exception:
        pass
    return ""

def clean_kecamatan_input(kec_raw: str) -> str:
    k = str(kec_raw or "").strip()
    if not k: return ""
    low = k.lower()
    if low in ["stm hilir","s.t.m hilir","st m hilir","stmhilir"]:
        return "SINEMBAH TANJUNG MUDA HILIR"
    if low in ["stm hulu","s.t.m hulu","st m hulu","stmhulu"]:
        return "SINEMBAH TANJUNG MUDA HULU"
    up = _norm_text(k)
    # Normalisasi variasi ke bentuk mapping "BIRU BIRU"
    if "SIBIRU BIRU" in up or "SI BIRU" in up:
        up = "BIRU BIRU"
    return up

def ui_kecamatan_input(kec_raw: str) -> str:
    """Nilai yang DIKETIK ke UI untuk field Kecamatan.
       Khusus SIBIRU-BIRU → ketik 'BIRU' (bukan 'BIRU BIRU')."""
    v = clean_kecamatan_input(kec_raw)
    if "BIRU BIRU" in v:
        return "BIRU"
    return v

def clean_kelurahan_input(vill_raw: str) -> str:
    s = str(vill_raw or "").strip()
    if not s: return ""
    up = _norm_text(s)
    up = up.replace("DESA KEL", " ").replace("DESA/KEL", " ")
    up = up.replace("DESA", " ").replace("KELURAHAN", " ").replace("KEL", " ").replace("KEL.", " ")
    up = re.sub(r"\s+", " ", up).strip()
    return up

# ===== BACA REFERENSI List_Kecamatan.xlsx =====
def _find_col(df, keywords):
    for c in df.columns:
        n = _norm_text(c)
        if any(kw in n for kw in keywords):
            return c
    return None

try:
    df_ref = pd.read_excel(EXCEL_PATH_KEC, sheet_name=0, dtype=str)
except Exception as e:
    print(f"❌ Gagal baca {EXCEL_PATH_KEC}: {e}")
    sys.exit(1)

col_kec = _find_col(df_ref, ["KEC"]) or _find_col(df_ref, ["KECAMATAN"])
col_kab = _find_col(df_ref, ["KAB", "KOTA"])
col_prov= _find_col(df_ref, ["PROV"])
if not all([col_kec, col_kab, col_prov]):
    print("❌ List_Kecamatan.xlsx harus punya kolom 'Kecamatan', 'Kabupaten/Kota', dan 'Provinsi'.")
    sys.exit(1)

kec_to_region = {}
for _, r in df_ref.iterrows():
    kec_val  = (r.get(col_kec)  or "").strip()
    kab_val  = (r.get(col_kab)  or "").strip()
    prov_val = (r.get(col_prov) or "").strip()
    if not kec_val:
        continue
    base = _norm_text(kec_val)
    for k in {base, _compact(kec_val), base.replace(" ", "")}:
        kec_to_region[k] = (kab_val, prov_val)

def lookup_region(kec_raw: str):
    if not kec_raw:
        return ("", "")
    normalized = clean_kecamatan_input(kec_raw)
    keys = [normalized, normalized.replace(" ", ""), _compact(kec_raw)]
    for k in keys:
        if k in kec_to_region:
            return kec_to_region[k]
    for k in kec_to_region:
        if k in normalized or normalized in k:
            return kec_to_region[k]
    return ("", "")

# ====== DETEKSI HEADER FILE SISWA + INDEX KOLUM ======
wb = load_workbook(EXCEL_PATH_SISWA)
sheet = wb.active

def _norm_cell(v): return re.sub(r"\s+"," ",str(v or "").strip()).lower()

header_row, header_vals = None, None
for r in range(1, 16):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column+1)]
    nn = [_norm_cell(v) for v in vals]
    if any("nama" in x for x in nn) and (any(x=="jk" for x in nn) or any("l p" in x for x in nn) or any("l/p" in x for x in nn)) and any("tanggal lahir" in x for x in nn):
        header_row, header_vals = r, vals
        break
if header_row is None:
    header_row, header_vals = 1, [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column+1)]

print(f"ℹ️ Header terdeteksi di baris Excel: {header_row}")

name_to_idx = {_norm_cell(h): i for i, h in enumerate(header_vals)}

def find_col(*aliases):
    for a in aliases:
        a = _norm_cell(a)
        if a in name_to_idx: return name_to_idx[a]
    for a in aliases:
        a = _norm_cell(a)
        for k,i in name_to_idx.items():
            if a in k: return i
    return None

IDX_NAMA      = find_col("nama","nama siswa","nama peserta","nama lengkap")
IDX_JK        = find_col("jk","l p","l/p","jenis kelamin","kelamin")
IDX_TMP_LAHIR = find_col("tempat lahir","tmpt lahir","kota lahir")
IDX_TGL_LAHIR = find_col("tanggal lahir","tgl lahir","lahir")
IDX_NIK       = find_col("nik","no ktp","nomor ktp","no. ktp","no nik","no. nik")
IDX_AGAMA     = find_col("agama")
IDX_ALAMAT    = find_col("alamat","alamat domisili","alamat rumah")
IDX_KEC       = find_col("kecamatan","kec","kecamatan domisili")
IDX_KEL       = find_col("kelurahan","desa","desa/kel","desa/kelurahan","desa/kel.")

print("ℹ️ Index kolom:", {"nama":IDX_NAMA,"jk":IDX_JK,"tempat_lahir":IDX_TMP_LAHIR,"tanggal_lahir":IDX_TGL_LAHIR,"nik":IDX_NIK,"agama":IDX_AGAMA,"alamat":IDX_ALAMAT,"kecamatan":IDX_KEC,"kelurahan":IDX_KEL})

DATA_FIRST_ROW = header_row + 1

# ====== BROWSER ======
opt = Options()
opt.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=opt)
driver.maximize_window()
driver.get(URL_LOGIN)
print("✅ Silakan login manual (captcha) lalu klik LOGIN")
input("🔐 Tekan ENTER setelah BERHASIL login...")

# buka halaman pasien
driver.get(URL_PASIEN)
time.sleep(2)

# ===== Helper autocomplete wilayah (dalam modal) =====
def select_autocomplete_field(label_text, value, prefix=""):
    if not value:
        return
    try:
        xpath_input = f'//div[@id="modalTambahData"]//label[contains(text(), "{label_text}")]/following-sibling::div//input'
        for _ in range(3):  # retry max 3x
            try:
                input_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, xpath_input))
                )
                input_field.click(); time.sleep(0.3)
                input_field.clear(); time.sleep(0.3)
                input_field.send_keys(f"{prefix}{value}".strip())
                time.sleep(1.0)
                input_field.send_keys(Keys.ARROW_DOWN)
                input_field.send_keys(Keys.ENTER)
                time.sleep(1.0)
                return
            except StaleElementReferenceException:
                time.sleep(1.0)
        print(f"⚠️ Gagal set field '{label_text}' dengan nilai '{value}'")
    except Exception as e:
        print(f"⚠️ Error saat isi '{label_text}': {e}")

# ====== LOOP ======
FAILED = []
JUMLAH_DATA = NOMOR_SISWA_AKHIR - NOMOR_SISWA_AWAL + 1

for i in range(JUMLAH_DATA):
    excel_row = DATA_FIRST_ROW + (NOMOR_SISWA_AWAL - 1) + i
    row = sheet[excel_row]

    def get(idx): return None if idx is None else row[idx].value

    nama         = get(IDX_NAMA)
    if not nama:
        print(f"⚠️ Baris {excel_row}: nama kosong. Skip.")
        continue
    jk           = str(get(IDX_JK) or "").strip()
    tempat_lahir = get(IDX_TMP_LAHIR)
    tgl_raw      = get(IDX_TGL_LAHIR)
    nik          = str(get(IDX_NIK) or "")
    agama        = str(get(IDX_AGAMA) or "").strip()
    alamat       = get(IDX_ALAMAT)
    kec_raw      = get(IDX_KEC)
    kel_raw      = get(IDX_KEL)

    tanggal_lahir = fmt_date(tgl_raw, OUT_DATE_FMT)
    if not tanggal_lahir:
        print(f"⚠️ Gagal format tanggal: {nama} -> {tgl_raw}")

    kabupaten_disp, provinsi_disp = lookup_region(kec_raw)
    kelurahan_clean = clean_kelurahan_input(kel_raw)
    kec_ui_value    = ui_kecamatan_input(kec_raw)

    print(f"\n▶ {i+1}. {nama} | Kec='{kec_raw}' → Kab='{kabupaten_disp}' | Prov='{provinsi_disp}' | Kel='{kelurahan_clean}'")
    input("🟢 Klik tombol 'TAMBAH DATA' untuk membuka form. Setelah modal terbuka, tekan ENTER untuk mengisi otomatis...")

    try:
        # Pastikan modal sudah terbuka
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "modalTambahData")))

        # === Identitas ===
        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Nama"]').clear()
        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Nama"]').send_keys(str(nama))

        Select(driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//select[./option[contains(text(), "Jenis Kelamin")]]')) \
            .select_by_visible_text("Perempuan" if jk.upper() == "P" else "Laki-Laki")

        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Tempat Lahir"]').clear()
        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Tempat Lahir"]').send_keys(str(tempat_lahir or ""))

        el = driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Tanggal Lahir"]')
        el.clear(); el.send_keys(tanggal_lahir)

        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="NIK"]').clear()
        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="NIK"]').send_keys(nik)

        if agama:
            Select(driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//select[./option[contains(text(), "Agama")]]')) \
                .select_by_visible_text(agama.upper())

        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Alamat"]').clear()
        driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Alamat"]').send_keys(str(alamat or ""))

        # === Wilayah ===
        if provinsi_disp:
            select_autocomplete_field("Provinsi", provinsi_disp)

        if kabupaten_disp:
            kab_text = kabupaten_disp.strip()
            if not kab_text.upper().startswith(("KABUPATEN ", "KOTA ")):
                kab_text = "KABUPATEN " + kab_text
            select_autocomplete_field("Kabupaten", kab_text)

        if kec_ui_value:
            select_autocomplete_field("Kecamatan", kec_ui_value)

        if kelurahan_clean:
            select_autocomplete_field("Kelurahan", kelurahan_clean)

        print("✅ Data berhasil diisi.")

    except Exception as e:
        print(f"❌ Gagal input untuk {nama} → {e}")
        FAILED.append(str(nama))
        # Coba tutup modal (biar siap untuk iterasi berikutnya)
        try:
            close_btn = driver.find_element(By.XPATH, '//div[@id="modalTambahData"]//button[contains(., "CLOSE") or contains(., "Close")]')
            close_btn.click()
            WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, "modalTambahData")))
        except Exception:
            pass
        continue

# ===== TULIS LOG GAGAL =====
try:
    with open(FAILED_LOG_PATH, "w", encoding="utf-8") as f:
        f.write(f"TOTAL YANG GAGAL INPUT: {len(FAILED)}\n")
        for nm in FAILED:
            f.write(f"{nm}\n")
    print(f"\n📝 Log kegagalan ditulis ke '{FAILED_LOG_PATH}'.")
except Exception as e:
    print(f"\n⚠️ Gagal menulis '{FAILED_LOG_PATH}': {e}")

print("\n🍀 Selesai semua baris.")
