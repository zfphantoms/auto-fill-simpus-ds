# auto_input_simpus_fix.py
# Mode MANUAL:
# - Kamu klik "TAMBAH DATA" sendiri (script menunggu ENTER lalu isi otomatis)
# - Setelah terisi, kamu klik "TAMBAH" sendiri (script menunggu ENTER baru lanjut)
# Fitur:
# - Wilayah resolve via wilayah_indonesia.csv (Provinsi, Kab/Kota, Kecamatan, Desa)
# - Autocomplete exact-match selection (choose_mui_autocomplete)
# - Tanggal lahir via JavaScript (timezone-safe, lintas locale)
# - No KK opsional (jika ada kolom di Excel)
# - No BPJS / JKN / KIS opsional (jika ada kolom di Excel)
# - Status Kawin opsional (jika ada kolom di Excel)
# - Log gagal ke failed-log.txt
# - Khusus SIBIRU-BIRU: yang DIKETIK ke UI = "BIRU" (mapping tetap "BIRU-BIRU")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException

# Explicit imports to help PyInstaller bundle dynamic dependencies
import selenium.webdriver.chrome.webdriver
import selenium.webdriver.chrome.service
import selenium.webdriver.chrome.options
from openpyxl import load_workbook
from datetime import datetime
import time, re, sys
import pandas as pd

from wilayah_loader import load_wilayah_csv, build_wilayah_index, resolve_wilayah

# === KONFIGURASI ===
URL_LOGIN = "https://dinkesds-simpus.deliserdangkab.go.id/"
URL_PASIEN = "https://dinkesds-simpus.deliserdangkab.go.id/pasien"
WILAYAH_CSV = "wilayah_indonesia.csv"
FAILED_LOG_PATH = "failed-log.txt"

# === PILIH FILE EXCEL (via File Explorer) ===
import tkinter as tk
from tkinter import filedialog

_root = tk.Tk()
_root.withdraw()  # Sembunyikan jendela utama tkinter
_root.attributes('-topmost', True)  # Pastikan dialog muncul di depan

print("\n📂 Silakan pilih file Excel di jendela yang muncul...")
EXCEL_PATH_SISWA = filedialog.askopenfilename(
    title="Pilih file Excel data siswa",
    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    initialdir="."
)
_root.destroy()

if not EXCEL_PATH_SISWA:
    print("❌ Tidak ada file yang dipilih. Keluar.")
    sys.exit(1)

print(f"✅ File yang dipilih: {EXCEL_PATH_SISWA}\n")

# === INPUT USER ===
NOMOR_SISWA_AWAL = int(input("Masukkan nomor urut siswa awal (1-based): "))
NOMOR_SISWA_AKHIR = int(input("Masukkan nomor urut siswa akhir: "))
if NOMOR_SISWA_AKHIR < NOMOR_SISWA_AWAL:
    raise ValueError("❌ Nomor siswa akhir harus >= awal.")


# =====================================================================
# UTIL: Normalisasi (diadopsi dari open_simpus.py)
# =====================================================================

def _norm(s: str) -> str:
    """Normalisasi ringan: lowercase, trim, collapse spasi."""
    return ' '.join((s or '').strip().lower().split())


def _norm_cell(v):
    """Normalisasi header/cell Excel: lowercase, trim, collapse spasi."""
    return re.sub(r"\s+", " ", str(v or "").strip()).lower()


def _clean_kecamatan_input_for_lookup(kec_raw: str) -> str:
    """Normalisasi kecamatan untuk proses lookup kamus.

    Mengadopsi kasus-kasus ambigu (STM, SIBIRU-BIRU, dll).
    Juga membuang prefix 'Kec.', 'Kec', 'Kecamatan'.
    """
    k = str(kec_raw or '').strip()
    if not k:
        return ''

    # Buang prefix Kec. / Kec / Kecamatan
    k = re.sub(r'^(?:KECAMATAN|KEC\.?\s*)\s+', '', k, flags=re.IGNORECASE).strip()

    low = k.lower().strip()

    # Variasi STM Hilir/Hulu
    if low in {'stm hilir', 's.t.m hilir', 'st m hilir', 'stmhilir'}:
        return 'SINEMBAH TANJUNG MUDA HILIR'
    if low in {'stm hulu', 's.t.m hulu', 'st m hulu', 'stmhulu'}:
        return 'SINEMBAH TANJUNG MUDA HULU'

    # Variasi SIBIRU-BIRU
    up = str(k).strip().upper().replace('-', ' ')
    up = re.sub(r"\s+", " ", up).strip()
    if 'SIBIRU BIRU' in up or 'SI BIRU' in up:
        return 'BIRU BIRU'

    return k


def _clean_desa_kelurahan_for_lookup(name: str) -> str:
    """Normalisasi desa/kelurahan untuk lookup kamus.

    - buang kata DESA/KEL/KELURAHAN/DS.
    - buang tanda baca umum
    - rapikan spasi
    """
    s = str(name or '').strip()
    if not s:
        return ''

    # Alias spesifik untuk kasus yang sering muncul di sumber data
    alias_map = {
        'KELAMBIR LIMA KEBUN': 'KLAMBIR LIMA KEBUN',
    }
    up0 = re.sub(r"\s+", " ", s.replace('\u00a0', ' ')).strip().upper()
    if up0 in alias_map:
        s = alias_map[up0]

    # rapikan NBSP
    s = s.replace('\u00a0', ' ')

    # buang prefix yang sering muncul
    s = re.sub(r"\bDESA\s*KEL\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bDESA/KEL\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bDESA\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bKELURAHAN\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bKEL\.?\b", " ", s, flags=re.IGNORECASE)

    # ganti '-' jadi spasi, buang karakter aneh
    s = s.replace('-', ' ')
    s = re.sub(r"[^0-9a-zA-Z ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s


def _kecamatan_value_for_ui(kec_raw_or_resolved: str) -> str:
    """Nilai yang diketik ke UI untuk field Kecamatan.

    Khusus SIBIRU-BIRU: UI lebih mudah jika diketik 'BIRU' (bukan 'BIRU BIRU').
    """
    v = str(_clean_kecamatan_input_for_lookup(kec_raw_or_resolved) or '').strip()
    if not v:
        return v

    up = v.upper()
    if 'BIRU BIRU' in up:
        return 'BIRU'

    return v


def _normalize_provinsi_for_ui(name: str) -> str:
    """Normalize province name for SIMPUS autocomplete."""
    s = (name or '').strip().strip('"').strip("'").strip()
    if not s:
        return s

    up = s.upper().strip()

    # Common aliases
    if up in {'KEPRI', 'KEP. RIAU', 'KEPULAUAN RIAU'}:
        return 'KEPULAUAN RIAU'

    # Ensure 'RIAU' remains exactly 'RIAU'
    if up == 'RIAU':
        return 'RIAU'

    return s


def _strip_admin_prefix_kab_kota(name: str) -> str:
    """Return core Kab/Kota name without leading admin prefixes.

    Examples:
    - 'KAB. DELI SERDANG' -> 'DELI SERDANG'
    - 'KABUPATEN DELI SERDANG' -> 'DELI SERDANG'
    - 'KOTA MEDAN' -> 'MEDAN'
    """
    s = (name or '').strip()
    if not s:
        return s

    up = s.upper().strip()
    prefixes = [
        'KABUPATEN ',
        'KAB. ',
        'KAB ',
        'KOTA ADMINISTRASI ',
        'KOTA ',
    ]
    for p in prefixes:
        if up.startswith(p):
            return s[len(p):].strip()
    return s


def _normalize_desa_kelurahan_for_ui(name: str) -> str:
    """Normalize desa/kelurahan name for SIMPUS autocomplete.

    Handles known spelling variations.
    """
    s = (name or '').strip()
    if not s:
        return s

    up = s.upper().strip()

    # Known aliases (extend as needed)
    aliases = {
        'KLUMPANG KEBUN': 'KLUMPANG KEBON',
        'KLAMBIR LIMA KEBUN': 'KLAMBIR LIMA KEBON',
        'SEI BAHARU': 'SUNGAI BAHARU',
    }
    if up in aliases:
        return aliases[up]

    # Kasus khusus: TANDEM/TANDAM
    if up.startswith('TANDEM ') or up.startswith('TANDAM '):
        base = up.replace('TANDEM', 'TANDAM', 1)
        base = re.sub(r"\bI\b", "SATU", base)
        base = re.sub(r"\bII\b", "DUA", base)
        base = re.sub(r"\bIII\b", "TIGA", base)
        if base == 'TANDAM HULU SATU':
            return 'KAMPUNG TANDAM HULU SATU'
        return base

    return s


def _norm_wilayah_key(s: str) -> str:
    """Normalisasi ringan untuk kecamatan/desa sebelum lookup kamus."""
    s = (s or '').strip()
    if not s:
        return s
    s = s.replace('\u00a0', ' ')
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"^(KEL\.?|KELURAHAN|DESA)\s+", "", s, flags=re.IGNORECASE).strip()
    return s


# --- Numeric variant expansion (Romawi <-> kata) ---
_ROMAN_TO_INT = {
    'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5,
    'vi': 6, 'vii': 7, 'viii': 8, 'ix': 9, 'x': 10,
}
_INTWORD_TO_INT = {
    'satu': 1, 'dua': 2, 'tiga': 3, 'empat': 4, 'lima': 5,
    'enam': 6, 'tujuh': 7, 'delapan': 8, 'sembilan': 9, 'sepuluh': 10,
}


def _expand_numeric_variants(name: str) -> list:
    """Buat variasi nama dengan konversi angka romawi <-> kata.

    Contoh: 'KELAMBIR LIMA KEBUN' <-> 'KELAMBIR V KEBUN'.
    """
    base = _norm_wilayah_key(name)
    if not base:
        return []

    tokens = base.split()
    variants = {base}

    # roman -> kata
    for i, t in enumerate(tokens):
        t_clean = re.sub(r"[^0-9a-zA-Z]", "", t).lower()
        if t_clean in _ROMAN_TO_INT:
            n = _ROMAN_TO_INT[t_clean]
            word = next((w for w, nn in _INTWORD_TO_INT.items() if nn == n), None)
            if word:
                new_tokens = tokens[:]
                new_tokens[i] = word
                variants.add(' '.join(new_tokens))

    # kata -> roman
    for i, t in enumerate(tokens):
        t_clean = re.sub(r"[^0-9a-zA-Z]", "", t).lower()
        if t_clean in _INTWORD_TO_INT:
            n = _INTWORD_TO_INT[t_clean]
            roman = next((r for r, nn in _ROMAN_TO_INT.items() if nn == n), None)
            if roman:
                new_tokens = tokens[:]
                new_tokens[i] = roman.upper()
                variants.add(' '.join(new_tokens))

    return sorted(variants)


def resolve_wilayah_with_fallback(idx, kecamatan: str, desa: str):
    """Resolve wilayah dengan beberapa kandidat normalisasi.

    Urutan:
    1) coba apa adanya
    2) coba normalisasi prefix/spasi
    3) coba variasi angka (LIMA <-> V, dst)
    4) coba compact match (DELITUA == DELI TUA, hapus semua spasi)
    """
    kecamatan = _clean_kecamatan_input_for_lookup(kecamatan)
    desa = _clean_desa_kelurahan_for_lookup(desa)

    kec_raw = (kecamatan or '').strip()
    desa_raw = (desa or '').strip()

    # 1) raw
    cands = resolve_wilayah(idx, kec_raw, desa_raw)
    if cands:
        return cands

    # 2) normalisasi ringan
    kec_norm = _norm_wilayah_key(kec_raw)
    desa_norm = _norm_wilayah_key(desa_raw)
    cands = resolve_wilayah(idx, kec_norm, desa_norm)
    if cands:
        return cands

    # 2b) variasi desa/kel khusus
    if _norm(desa_norm) == 'kelambir':
        for alt in ['Dagang Kelambir']:
            cands = resolve_wilayah(idx, kec_norm or kec_raw, alt)
            if cands:
                return cands

    # 3) variasi numeric
    kec_variants = _expand_numeric_variants(kec_raw) or [kec_norm or kec_raw]
    desa_variants = _expand_numeric_variants(desa_raw) or [desa_norm or desa_raw]

    tried = set()
    for k in kec_variants:
        for d in desa_variants:
            kk = _norm_wilayah_key(k)
            dd = _norm_wilayah_key(d)
            key = (_norm(kk), _norm(dd))
            if key in tried:
                continue
            tried.add(key)
            cands = resolve_wilayah(idx, kk, dd)
            if cands:
                return cands

    # 4) Compact match: hapus semua spasi lalu bandingkan
    #    Menangani kasus "DELITUA" == "Deli Tua", "DELITUA BARAT" == "Deli Tua Barat"
    kec_compact = _norm(kec_norm or kec_raw).replace(' ', '')
    desa_compact = _norm(desa_norm or desa_raw).replace(' ', '')

    for (idx_kec, idx_desa), records in idx.items():
        kec_match = (idx_kec.replace(' ', '') == kec_compact) if kec_compact else (not idx_kec)
        desa_match = (idx_desa.replace(' ', '') == desa_compact) if desa_compact else (not idx_desa)
        if kec_match and desa_match:
            return records

    return []


# =====================================================================
# UTIL: Tanggal (JavaScript-based, timezone-safe)
# =====================================================================

def normalize_date_yyyy_mm_dd(val):
    """Konversi value tanggal (dari openpyxl atau string) ke 'YYYY-MM-DD'."""
    if val is None:
        return None
    # Jika sudah datetime object (dari openpyxl)
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    # Potong jika ada waktu (mis. '2000-01-15 00:00:00')
    s_date = s[:10] if len(s) > 10 else s
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            d = datetime.strptime(s_date, fmt)
            return d.strftime("%Y-%m-%d")
        except Exception:
            continue
    # Fallback dengan pandas
    try:
        d = pd.to_datetime(s, dayfirst=True, errors="raise")
        return d.strftime("%Y-%m-%d")
    except Exception:
        return None


def set_date_input_js(driver_ref, date_input, yyyy_mm_dd: str):
    """Set <input type=date> secara stabil lintas format (mm/dd vs dd/mm) menggunakan JS.

    Untuk mencegah off-by-one karena timezone, set pakai UTC tengah hari.
    """
    driver_ref.execute_script(
        """
        const el = arguments[0];
        const v = arguments[1];
        if (!v) return;
        const parts = v.split('-');
        const y = parseInt(parts[0], 10);
        const m = parseInt(parts[1], 10) - 1;
        const d = parseInt(parts[2], 10);

        // Buat tanggal pada UTC tengah hari untuk menghindari pergeseran timezone
        const utcDate = new Date(Date.UTC(y, m, d, 12, 0, 0));
        el.valueAsDate = utcDate;

        // fallback kalau browser tidak konsisten dengan valueAsDate
        if (el.value !== v) {
            el.value = v;
        }

        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        date_input,
        yyyy_mm_dd,
    )


# =====================================================================
# UTIL: Gender & Status Kawin mapping
# =====================================================================

def map_gender_to_option(value: str):
    """Mapping jenis kelamin dari Excel ke opsi dropdown SIMPUS."""
    v = (value or '').strip().lower()
    if not v or v == 'nan':
        return None
    if v in ['1', 'l', 'lk', 'laki', 'laki-laki', 'male']:
        return 'Laki-Laki'
    if v in ['2', 'p', 'pr', 'perempuan', 'female', 'wanita']:
        return 'Perempuan'
    if 'laki' in v:
        return 'Laki-Laki'
    if 'perempuan' in v or 'wanita' in v:
        return 'Perempuan'
    return None


def map_status_kawin_to_option(value: str):
    """Mapping status kawin dari Excel ke opsi dropdown SIMPUS.

    Opsi di form SIMPUS (dropdown 'Status Hubungan'):
    - Lajang
    - Kawin
    - Janda
    - Duda
    - Lainnya
    """
    v = (value or '').strip().lower()
    if not v or v == 'nan':
        return None
    # Angka: 1=Lajang, 2=Kawin
    if v == '1':
        return 'Lajang'
    if v == '2':
        return 'Kawin'
    # Teks exact/alias
    if v in ['lajang', 'single', 'belum kawin', 'belum menikah', 'belum nikah', 'bk']:
        return 'Lajang'
    if v in ['kawin', 'menikah', 'nikah', 'k', 'married', 'sudah kawin', 'sudah menikah']:
        return 'Kawin'
    if v in ['janda']:
        return 'Janda'
    if v in ['duda']:
        return 'Duda'
    # Partial match
    if 'belum' in v:
        return 'Lajang'
    if 'janda' in v:
        return 'Janda'
    if 'duda' in v:
        return 'Duda'
    if 'kawin' in v or 'nikah' in v or 'menikah' in v:
        return 'Kawin'
    # Fallback
    return 'Lainnya'


def clean_digits(value):
    """Bersihkan nilai menjadi hanya digit (untuk NIK, No KK, No BPJS)."""
    s = str(value or '').strip()
    if not s or s.lower() == 'nan':
        return None
    s = ''.join(ch for ch in s if ch.isdigit())
    return s if s else None


# =====================================================================
# LOAD WILAYAH
# =====================================================================

print("ℹ️ Memuat referensi wilayah dari wilayah_indonesia.csv...")
try:
    _records = load_wilayah_csv(WILAYAH_CSV)
    WILAYAH_IDX = build_wilayah_index(_records)
    print(f"ℹ️ Wilayah dimuat: {len(_records)} record.")
except Exception as e:
    print(f"❌ Gagal memuat {WILAYAH_CSV}: {e}")
    sys.exit(1)


# =====================================================================
# DETEKSI HEADER FILE SISWA + INDEX KOLOM
# =====================================================================

wb = load_workbook(EXCEL_PATH_SISWA)
sheet = wb.active

header_row, header_vals = None, None
for r in range(1, 16):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
    nn = [_norm_cell(v) for v in vals]
    if (any("nama" in x for x in nn)
            and (any(x == "jk" for x in nn) or any("l p" in x for x in nn) or any("l/p" in x for x in nn) or any("jenis kelamin" in x for x in nn))
            and any("tanggal lahir" in x or "tgl lahir" in x for x in nn)):
        header_row, header_vals = r, vals
        break
if header_row is None:
    header_row, header_vals = 1, [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]

print(f"ℹ️ Header terdeteksi di baris Excel: {header_row}")

name_to_idx = {_norm_cell(h): i for i, h in enumerate(header_vals)}


def find_col(*aliases):
    """Cari index kolom berdasarkan alias (exact match dulu, lalu partial)."""
    for a in aliases:
        a = _norm_cell(a)
        if a in name_to_idx:
            return name_to_idx[a]
    for a in aliases:
        a = _norm_cell(a)
        for k, i in name_to_idx.items():
            if a in k:
                return i
    return None


# --- Kolom wajib ---
IDX_NAMA      = find_col("nama", "nama siswa", "nama peserta", "nama lengkap")
IDX_JK        = find_col("jk", "l p", "l/p", "jenis kelamin", "kelamin")
IDX_TMP_LAHIR = find_col("tempat lahir", "tmpt lahir", "kota lahir")
IDX_TGL_LAHIR = find_col("tanggal lahir", "tgl lahir", "lahir")
IDX_NIK       = find_col("nik", "no ktp", "nomor ktp", "no. ktp", "no nik", "no. nik", "nik/kitas/kitap", "nomor induk kependudukan")
IDX_NO_KK     = find_col("no kk", "no. kk", "nomor kk", "kk", "nomor kartu keluarga", "kartu keluarga", "no kartu keluarga", "no. kartu keluarga", "no.kk")
IDX_AGAMA     = find_col("agama")
IDX_ALAMAT    = find_col("alamat", "alamat domisili", "alamat rumah", "alamat tempat tinggal", "alamat lengkap")
IDX_KEC       = find_col("kecamatan", "kec", "kecamatan domisili", "nama kecamatan")
IDX_KEL       = find_col("kelurahan", "desa", "desa/kel", "desa/kelurahan", "desa/kel.", "nama desa", "desa kelurahan", "kelurahan desa")

# --- Kolom opsional ---
IDX_BPJS          = find_col("no bpjs", "no. bpjs", "bpjs", "nomor bpjs", "no jkn", "jkn", "no. jkn", "nomor jkn", "no kis", "kis", "no. kis", "nomor kis")
IDX_STATUS_KAWIN  = find_col("status kawin", "status perkawinan", "kawin", "perkawinan", "status nikah")
IDX_RT            = find_col("rt")
IDX_RW            = find_col("rw")

print("ℹ️ Index kolom (wajib):", {
    "nama": IDX_NAMA, "jk": IDX_JK, "tempat_lahir": IDX_TMP_LAHIR,
    "tanggal_lahir": IDX_TGL_LAHIR, "nik": IDX_NIK, "no_kk": IDX_NO_KK,
    "agama": IDX_AGAMA, "alamat": IDX_ALAMAT, "kecamatan": IDX_KEC, "kelurahan": IDX_KEL
})
print("ℹ️ Index kolom (opsional):", {
    "bpjs": IDX_BPJS, "status_kawin": IDX_STATUS_KAWIN, "rt": IDX_RT, "rw": IDX_RW
})
if IDX_BPJS is not None:
    print("  ⏭️ Kolom No BPJS/JKN/KIS ditemukan namun fitur pengisian di-nonaktifkan → dilewati")
else:
    print("  ⏭️ Kolom No BPJS/JKN/KIS tidak ditemukan → dilewati")
if IDX_STATUS_KAWIN is not None:
    print("  ✅ Kolom Status Kawin ditemukan → akan diisi")
else:
    print("  ⏭️ Kolom Status Kawin tidak ditemukan → dilewati")
if IDX_RT is not None:
    print("  ✅ Kolom RT ditemukan → akan diisi")
else:
    print("  ⏭️ Kolom RT tidak ditemukan → dilewati")
if IDX_RW is not None:
    print("  ✅ Kolom RW ditemukan → akan diisi")
else:
    print("  ⏭️ Kolom RW tidak ditemukan → dilewati")

DATA_FIRST_ROW = header_row + 1


# =====================================================================
# BROWSER
# =====================================================================

opt = Options()
opt.add_experimental_option("detach", True)

print("\nMenyiapkan ChromeDriver (via built-in Selenium Manager)...")
# Gunakan bawaan Selenium Manager (lebih aman saat di-build jadi .exe)
driver = webdriver.Chrome(options=opt)
wait = WebDriverWait(driver, 20)
driver.maximize_window()
driver.get(URL_LOGIN)

print("✅ Silakan login manual (captcha) lalu klik LOGIN")
print("[INFO] Menunggu login terdeteksi secara otomatis...")

# Deteksi login otomatis: tunggu sampai URL berubah ke /home
timeout_seconds = 300  # 5 menit
poll_interval_seconds = 1
start = time.time()

while True:
    current_url = driver.current_url or ''
    if '/home' in current_url:
        break
    if time.time() - start > timeout_seconds:
        print("❌ Timeout: belum terdeteksi login (belum masuk /home).")
        driver.quit()
        sys.exit(1)
    time.sleep(poll_interval_seconds)

print("[INFO] Login terdeteksi! Mengalihkan ke halaman pasien...")

# buka halaman pasien
driver.get(URL_PASIEN)
time.sleep(2)


# =====================================================================
# Helper: choose_mui_autocomplete (dari open_simpus.py)
# =====================================================================

def choose_mui_autocomplete(label_text: str, value_to_choose: str):
    """Pilih nilai pada MUI Autocomplete di dalam modal #modalTambahData.

    Strategi:
    - Cari input berdasarkan label text
    - Ketik value
    - Jika list option muncul: pilih yang *exact match* (case-insensitive)
      Kalau tidak ada exact, fallback ARROWDOWN+ENTER.
    """
    if not value_to_choose:
        return

    modal = wait.until(EC.presence_of_element_located((By.ID, "modalTambahData")))

    xpath_input = (
        f".//label[contains(normalize-space(), '{label_text}')]"
        f"/following-sibling::div//input[@role='combobox' or @type='text' or @type='search' or not(@type)]"
    )

    target_norm = _norm(value_to_choose)

    last_err = None
    for _ in range(3):
        try:
            input_el = WebDriverWait(modal, 10).until(
                EC.element_to_be_clickable((By.XPATH, xpath_input))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_el)
            input_el.click()
            time.sleep(0.2)
            input_el.clear()
            time.sleep(0.2)
            input_el.send_keys(value_to_choose)
            time.sleep(0.6)

            # Coba pilih exact match dari listbox jika ada
            try:
                listbox = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.XPATH, "//ul[@role='listbox']"))
                )
                options = listbox.find_elements(By.XPATH, ".//li[@role='option']")
                exact = [o for o in options if _norm(o.text) == target_norm]
                if exact:
                    exact[0].click()
                    time.sleep(0.4)
                    return
            except Exception:
                # listbox tidak muncul cepat, lanjut fallback
                pass

            # Fallback: pilih item pertama
            input_el.send_keys(Keys.ARROW_DOWN)
            input_el.send_keys(Keys.ENTER)
            time.sleep(0.6)
            return
        except StaleElementReferenceException as e:
            last_err = e
            time.sleep(0.5)
        except Exception as e:
            last_err = e
            time.sleep(0.5)

    print(f"⚠️ Gagal set MUI Autocomplete '{label_text}' -> '{value_to_choose}': {last_err}")


def open_tambah_data_modal():
    """Klik tombol TAMBAH DATA dan tunggu modal terbuka."""
    tambah_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH,
        "//button[contains(@class,'btn') and contains(@data-bs-target,'#modalTambahData') and normalize-space()='TAMBAH DATA']"
    )))
    tambah_btn.click()
    wait.until(EC.visibility_of_element_located((By.ID, 'modalTambahData')))


def submit_tambah_and_wait_close():
    """Klik tombol submit TAMBAH di dalam modal dan tunggu modal tertutup."""
    submit_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH,
        "//div[@id='modalTambahData']//button[@type='submit' and contains(@class,'btn') and normalize-space()='TAMBAH']"
    )))
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", submit_btn)
    submit_btn.click()
    # Tunggu modal tertutup
    wait.until(EC.invisibility_of_element_located((By.ID, 'modalTambahData')))
    time.sleep(0.5)


def close_modal_if_open():
    """Tutup modal jika masih terbuka (untuk recovery dari error)."""
    try:
        modal = driver.find_element(By.ID, 'modalTambahData')
        if not modal.is_displayed():
            return
    except Exception:
        return
    for xp in [
        "//div[@id='modalTambahData']//button[contains(.,'CLOSE') or contains(.,'Close') or contains(.,'TUTUP') or contains(.,'Tutup')]",
        "//div[@id='modalTambahData']//button[contains(@class,'btn-close') or @aria-label='Close']",
    ]:
        try:
            btn = driver.find_element(By.XPATH, xp)
            btn.click()
            WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.ID, 'modalTambahData')))
            time.sleep(0.3)
            return
        except Exception:
            pass


# =====================================================================
# LOOP UTAMA
# =====================================================================

FAILED = []
success_count = 0
last_success_nama = None
last_success_nik = None
JUMLAH_DATA = NOMOR_SISWA_AKHIR - NOMOR_SISWA_AWAL + 1

for i in range(JUMLAH_DATA):
    excel_row = DATA_FIRST_ROW + (NOMOR_SISWA_AWAL - 1) + i
    row = sheet[excel_row]

    def get(idx):
        return None if idx is None else row[idx].value

    nama = get(IDX_NAMA)
    if not nama:
        print(f"⚠️ Baris {excel_row}: nama kosong. Skip.")
        continue

    jk = str(get(IDX_JK) or "").strip()
    tempat_lahir = str(get(IDX_TMP_LAHIR) or "").strip()
    tgl_raw = get(IDX_TGL_LAHIR)
    nik_raw = get(IDX_NIK)
    agama = str(get(IDX_AGAMA) or "").strip()
    alamat = str(get(IDX_ALAMAT) or "").strip()
    kec_raw = str(get(IDX_KEC) or "").strip()
    kel_raw = str(get(IDX_KEL) or "").strip()

    # Kolom wajib tambahan
    no_kk_raw = get(IDX_NO_KK)

    # Kolom opsional
    bpjs_raw = get(IDX_BPJS) if IDX_BPJS is not None else None
    status_kawin_raw = str(get(IDX_STATUS_KAWIN) or "").strip() if IDX_STATUS_KAWIN is not None else None
    rt_raw = get(IDX_RT) if IDX_RT is not None else None
    rw_raw = get(IDX_RW) if IDX_RW is not None else None

    # --- Proses data ---
    nama_value = str(nama).strip()
    gender_option = map_gender_to_option(jk)
    tanggal_lahir_value = normalize_date_yyyy_mm_dd(tgl_raw)
    nik_value = clean_digits(nik_raw)
    no_kk_value = clean_digits(no_kk_raw)
    bpjs_value = clean_digits(bpjs_raw)
    status_kawin_option = map_status_kawin_to_option(status_kawin_raw) if status_kawin_raw else None
    
    rt_value = str(rt_raw).strip() if rt_raw is not None and str(rt_raw).strip().lower() != 'nan' else None
    if rt_value and rt_value.endswith('.0'): rt_value = rt_value[:-2]
    
    rw_value = str(rw_raw).strip() if rw_raw is not None and str(rw_raw).strip().lower() != 'nan' else None
    if rw_value and rw_value.endswith('.0'): rw_value = rw_value[:-2]

    if not tanggal_lahir_value:
        print(f"⚠️ Gagal format tanggal: {nama_value} -> {tgl_raw}")

    # --- Resolve wilayah ---
    kec_cleaned = _clean_kecamatan_input_for_lookup(kec_raw)
    kel_cleaned = _clean_desa_kelurahan_for_lookup(kel_raw)

    candidates = resolve_wilayah_with_fallback(WILAYAH_IDX, kec_raw, kel_raw)

    prov_value_for_ui = ""
    kab_value_for_ui = ""
    kec_value_for_ui = ""
    kel_value_for_ui = ""

    if candidates:
        if len(candidates) > 1:
            preview = '; '.join([
                f"{c.provinsi} / {c.kabupaten_kota} / {c.kecamatan} / {c.desa_kelurahan}"
                for c in candidates[:5]
            ])
            print(f"⚠️ Wilayah ambigu ({len(candidates)} kandidat) untuk Kec='{kec_raw}', Desa='{kel_raw}'. Pakai kandidat pertama.")
            print(f"   Kandidat: {preview}")

        wilayah = candidates[0]
        prov_value_for_ui = _normalize_provinsi_for_ui(wilayah.provinsi)
        kab_value_for_ui = _strip_admin_prefix_kab_kota(wilayah.kabupaten_kota)
        kec_value_for_ui = _kecamatan_value_for_ui(wilayah.kecamatan)
        kel_value_for_ui = _normalize_desa_kelurahan_for_ui(wilayah.desa_kelurahan)
    else:
        print(f"⚠️ Wilayah tidak ditemukan untuk Kec='{kec_raw}', Desa='{kel_raw}'. Field wilayah akan dilewati.")

    # --- Log ---
    print(f"\n=== INPUT {i + 1}/{JUMLAH_DATA} ===")
    print(f"  Nama: {nama_value} | NIK: {nik_value or '-'}")
    print(f"  Prov={prov_value_for_ui} | Kab={kab_value_for_ui} | Kec={kec_value_for_ui} | Kel={kel_value_for_ui}")

    try:
        # Klik TAMBAH DATA otomatis
        open_tambah_data_modal()

        # === IDENTITAS ===

        # Nama
        nama_input = wait.until(EC.element_to_be_clickable((
            By.XPATH,
            "//div[@id='modalTambahData']//input[@type='text' and @placeholder='Nama' and (not(@disabled) or @disabled='false')]"
        )))
        nama_input.click(); nama_input.clear(); nama_input.send_keys(nama_value)

        # Jenis Kelamin
        if gender_option:
            gender_select_el = wait.until(EC.element_to_be_clickable((
                By.XPATH,
                "//div[@id='modalTambahData']//select[contains(@class,'form-control') and .//option[normalize-space()='Jenis Kelamin']]"
            )))
            Select(gender_select_el).select_by_visible_text(gender_option)

        # Tempat Lahir
        if tempat_lahir and tempat_lahir.lower() != 'nan':
            tempat_lahir_input = wait.until(EC.element_to_be_clickable((
                By.XPATH,
                "//div[@id='modalTambahData']//input[@type='text' and @placeholder='Tempat Lahir' and (not(@disabled) or @disabled='false')]"
            )))
            tempat_lahir_input.click(); tempat_lahir_input.clear(); tempat_lahir_input.send_keys(tempat_lahir)

        # Tanggal Lahir (via JavaScript)
        if tanggal_lahir_value:
            tanggal_lahir_input = wait.until(EC.presence_of_element_located((
                By.XPATH,
                "//div[@id='modalTambahData']//input[@type='date' and (@placeholder='Tanggal Lahir' or contains(@placeholder,'Tanggal'))]"
            )))
            set_date_input_js(driver, tanggal_lahir_input, tanggal_lahir_value)

        # NIK
        if nik_value:
            nik_input = wait.until(EC.element_to_be_clickable((
                By.XPATH,
                "//div[@id='modalTambahData']//input[@placeholder='NIK' and (@type='number' or @inputmode='numeric')]"
            )))
            nik_input.click(); nik_input.clear(); nik_input.send_keys(nik_value)

        # No. KK
        if no_kk_value:
            kk_input = wait.until(EC.element_to_be_clickable((
                By.XPATH,
                "//div[@id='modalTambahData']//input[@placeholder='No. KK' and (@type='number' or @inputmode='numeric')]"
            )))
            kk_input.click(); kk_input.clear(); kk_input.send_keys(no_kk_value)

        # Agama
        if agama and agama.lower() != 'nan':
            try:
                Select(driver.find_element(
                    By.XPATH, '//div[@id="modalTambahData"]//select[./option[contains(text(), "Agama")]]'
                )).select_by_visible_text(agama.upper())
            except Exception:
                print(f"  ⚠️ Gagal set Agama: '{agama}'")

        # Status Hubungan / Status Kawin (opsional)
        if status_kawin_option:
            try:
                status_kawin_select = driver.find_element(
                    By.XPATH,
                    "//div[@id='modalTambahData']//select[contains(@class,'form-control') and .//option[normalize-space()='Status Hubungan']]"
                )
                Select(status_kawin_select).select_by_visible_text(status_kawin_option)
            except Exception:
                print(f"  ⚠️ Gagal set Status Hubungan: '{status_kawin_option}'")

        # No. BPJS (opsional) - Di-comment sesuai permintaan agar tidak diisi otomatis
        # if bpjs_value:
        #     try:
        #         bpjs_input = wait.until(EC.element_to_be_clickable((
        #             By.XPATH,
        #             "//div[@id='modalTambahData']//input[@placeholder='No. BPJS' and @type='text']"
        #         )))
        #         bpjs_input.click(); bpjs_input.clear(); bpjs_input.send_keys(bpjs_value)
        #     except Exception:
        #         print(f"  ⚠️ Gagal isi No. BPJS (field tidak ditemukan di form)")

        # === ALAMAT ===
        if alamat and alamat.lower() != 'nan':
            try:
                alamat_input = wait.until(EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[@id='modalTambahData']//div[contains(@class,'col-md-12')][.//p[contains(@class,'example-form-small') and normalize-space()='Tulis alamat lengkap']]//input[@type='text' and @placeholder='Alamat' and contains(@class,'form-control')]"
                )))
                alamat_input.click(); alamat_input.clear(); alamat_input.send_keys(alamat)
            except Exception:
                # Fallback: cari input Alamat yang lebih generik
                try:
                    alamat_input = driver.find_element(
                        By.XPATH, '//div[@id="modalTambahData"]//input[@placeholder="Alamat"]'
                    )
                    alamat_input.clear(); alamat_input.send_keys(alamat)
                except Exception:
                    print(f"  ⚠️ Gagal isi Alamat")

        # RT
        if rt_value:
            try:
                rt_input = wait.until(EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[@id='modalTambahData']//input[@type='text' and @placeholder='RT' and contains(@class,'form-control')]"
                )))
                rt_input.click(); rt_input.clear(); rt_input.send_keys(rt_value)
            except Exception:
                print(f"  ⚠️ Gagal isi RT")

        # RW
        if rw_value:
            try:
                rw_input = wait.until(EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[@id='modalTambahData']//input[@type='text' and @placeholder='RW' and contains(@class,'form-control')]"
                )))
                rw_input.click(); rw_input.clear(); rw_input.send_keys(rw_value)
            except Exception:
                print(f"  ⚠️ Gagal isi RW")

        # === WILAYAH (via MUI Autocomplete, exact-match) ===
        if prov_value_for_ui:
            choose_mui_autocomplete('Provinsi', prov_value_for_ui)

        if kab_value_for_ui:
            choose_mui_autocomplete('Kabupaten / Kota', kab_value_for_ui)

        if kec_value_for_ui:
            choose_mui_autocomplete('Kecamatan', kec_value_for_ui)

        if kel_value_for_ui:
            choose_mui_autocomplete('Kelurahan', kel_value_for_ui)

        print("✅ Data berhasil diisi. Menekan TAMBAH...")

        # === SUBMIT TAMBAH ===
        submit_tambah_and_wait_close()
        success_count += 1
        last_success_nama = nama_value
        last_success_nik = nik_value
        print(f"✅ Sukses submit: {nama_value}")

    except Exception as e:
        print(f"❌ Gagal input untuk {nama_value} → {e}")
        FAILED.append(str(nama_value))
        close_modal_if_open()
        continue


# =====================================================================
# TULIS LOG GAGAL
# =====================================================================

print(f"\nSELESAI. Sukses: {success_count}, Gagal: {len(FAILED)}")

if last_success_nama:
    import os
    excel_filename = os.path.basename(EXCEL_PATH_SISWA)
    log_text = f"File Excel: {excel_filename}\nTerakhir Berhasil Diinput -> Nama: {last_success_nama} | NIK: {last_success_nik or '-'}\n"
    print(f"\n📌 {log_text.strip()}")
    try:
        with open("last-success-log.txt", "w", encoding="utf-8") as f:
            f.write(log_text)
        print("📝 Data terakhir berhasil disimpan ke 'last-success-log.txt'.")
    except Exception as e:
        print(f"⚠️ Gagal menyimpan log terakhir: {e}")

if FAILED:
    for nm in FAILED[:20]:
        print(f"  - gagal: {nm}")

try:
    with open(FAILED_LOG_PATH, "w", encoding="utf-8") as f:
        f.write(f"TOTAL YANG GAGAL INPUT: {len(FAILED)}\n")
        for nm in FAILED:
            f.write(f"{nm}\n")
    print(f"\n📝 Log kegagalan ditulis ke '{FAILED_LOG_PATH}'.")
except Exception as e:
    print(f"\n⚠️ Gagal menulis '{FAILED_LOG_PATH}': {e}")

print("\nTekan Ctrl+C untuk menutup browser dan keluar.")
try:
    while True:
        time.sleep(1)
except KeyboardInterrupt:
    print("\n[INFO] Menutup browser...")
    driver.quit()
